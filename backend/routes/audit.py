import csv
import json
import io
from datetime import datetime
from flask import Blueprint, request, jsonify, Response
from flask_jwt_extended import jwt_required, get_jwt_identity

from database import db
from models import User, AuditLog

audit_bp = Blueprint("audit", __name__)


def require_admin():
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    if not user or user.role.name != "admin":
        return None, (jsonify({"error": "Accès administrateur requis"}), 403)
    return user, None


def build_query(args):
    query = AuditLog.query

    if args.get("admin_user_id"):
        query = query.filter(AuditLog.admin_user_id == int(args["admin_user_id"]))
    if args.get("action_type"):
        query = query.filter(AuditLog.action_type == args["action_type"])
    if args.get("entity_type"):
        query = query.filter(AuditLog.entity_type == args["entity_type"])
    if args.get("date_from"):
        try:
            query = query.filter(AuditLog.timestamp >= datetime.fromisoformat(args["date_from"]))
        except ValueError:
            pass
    if args.get("date_to"):
        try:
            query = query.filter(AuditLog.timestamp <= datetime.fromisoformat(args["date_to"]))
        except ValueError:
            pass
    if args.get("search"):
        term = f"%{args['search']}%"
        query = query.filter(
            db.or_(
                AuditLog.entity_name.ilike(term),
                AuditLog.description.ilike(term),
                AuditLog.action_type.ilike(term),
            )
        )

    return query.order_by(AuditLog.timestamp.desc())


@audit_bp.route("/", methods=["GET"])
@jwt_required()
def get_audit_logs():
    admin, err = require_admin()
    if err:
        return err

    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))

    query = build_query(request.args)
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "logs": [log.to_dict() for log in paginated.items],
        "total": paginated.total,
        "page": page,
        "per_page": per_page,
        "pages": paginated.pages,
    }), 200


@audit_bp.route("/stats", methods=["GET"])
@jwt_required()
def get_stats():
    admin, err = require_admin()
    if err:
        return err

    total = AuditLog.query.count()
    by_action = db.session.query(AuditLog.action_type, db.func.count(AuditLog.id)).group_by(AuditLog.action_type).all()
    by_entity = db.session.query(AuditLog.entity_type, db.func.count(AuditLog.id)).group_by(AuditLog.entity_type).all()

    return jsonify({
        "total": total,
        "by_action": [{"action": a, "count": c} for a, c in by_action],
        "by_entity": [{"entity": e, "count": c} for e, c in by_entity],
    }), 200


@audit_bp.route("/export/csv", methods=["GET"])
@jwt_required()
def export_csv():
    admin, err = require_admin()
    if err:
        return err

    logs = build_query(request.args).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Utilisateur", "Action", "Entité", "ID Entité", "Nom Entité", "Description", "IP", "Horodatage"])
    for log in logs:
        writer.writerow([
            log.id,
            log.admin_user.username if log.admin_user else "",
            log.action_type, log.entity_type,
            log.entity_id or "", log.entity_name or "",
            log.description or "", log.ip_address or "",
            log.timestamp.isoformat() if log.timestamp else "",
        ])
    output.seek(0)
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=journal_audit.csv"})


@audit_bp.route("/export/json", methods=["GET"])
@jwt_required()
def export_json():
    admin, err = require_admin()
    if err:
        return err

    logs = build_query(request.args).all()
    data = json.dumps([log.to_dict() for log in logs], indent=2, default=str, ensure_ascii=False)
    return Response(data, mimetype="application/json",
                    headers={"Content-Disposition": "attachment; filename=journal_audit.json"})


@audit_bp.route("/export/xlsx", methods=["GET"])
@jwt_required()
def export_xlsx():
    admin, err = require_admin()
    if err:
        return err

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    logs = build_query(request.args).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Journal d'audit"

    headers = ["ID", "Utilisateur", "Action", "Entité", "ID Entité", "Nom Entité", "Description", "IP", "Horodatage"]
    hfill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
    hfont = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    for row_i, log in enumerate(logs, 2):
        ws.cell(row=row_i, column=1, value=log.id)
        ws.cell(row=row_i, column=2, value=log.admin_user.username if log.admin_user else "")
        ws.cell(row=row_i, column=3, value=log.action_type)
        ws.cell(row=row_i, column=4, value=log.entity_type)
        ws.cell(row=row_i, column=5, value=log.entity_id or "")
        ws.cell(row=row_i, column=6, value=log.entity_name or "")
        ws.cell(row=row_i, column=7, value=log.description or "")
        ws.cell(row=row_i, column=8, value=log.ip_address or "")
        ws.cell(row=row_i, column=9, value=log.timestamp.isoformat() if log.timestamp else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 50
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=journal_audit.xlsx"})


@audit_bp.route("/action-types", methods=["GET"])
@jwt_required()
def get_action_types():
    admin, err = require_admin()
    if err:
        return err
    types = db.session.query(AuditLog.action_type).distinct().all()
    return jsonify([t[0] for t in types]), 200


@audit_bp.route("/entity-types", methods=["GET"])
@jwt_required()
def get_entity_types():
    admin, err = require_admin()
    if err:
        return err
    types = db.session.query(AuditLog.entity_type).distinct().all()
    return jsonify([t[0] for t in types]), 200
